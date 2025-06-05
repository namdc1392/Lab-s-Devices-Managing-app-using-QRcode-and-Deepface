import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import sys
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook
import qrcode

# Lấy đường dẫn thư mục chứa file đang chạy
if getattr(sys, 'frozen', False):
    CURRENT_DIR = os.path.dirname(sys.executable)
else:
    CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))

#LIKE Đường dẫn đến các file dữ liệu
DATA_FILE = os.path.join(CURRENT_DIR, "danh_sach_vat_dung.json")
CATEGORY_FILE = os.path.join(CURRENT_DIR, "danh_muc_loai_thiet_bi.json")
QR_CODE_DIR = os.path.join(CURRENT_DIR, "qr_codes")
IMAGE_DIR = os.path.join(CURRENT_DIR, "device_images")

# Tạo thư mục nếu chưa tồn tại
os.makedirs(QR_CODE_DIR, exist_ok=True)
os.makedirs(IMAGE_DIR, exist_ok=True)

def tao_ma_qr(ma_vat_dung, ten_vat_dung):
    data = f"Mã: {ma_vat_dung}\nTên: {ten_vat_dung}"
    img = qrcode.make(data)
    img.save(os.path.join(QR_CODE_DIR, f"{ma_vat_dung}.png"))

def load_categories():
    try:
        if os.path.exists(CATEGORY_FILE):
            with open(CATEGORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except:
        pass
    
    default_categories = ["Máy tính", "Thiết bị mạng", "Thiết bị âm thanh", "Thiết bị trình chiếu"]
    try:
        with open(CATEGORY_FILE, "w", encoding="utf-8") as f:
            json.dump(default_categories, f, ensure_ascii=False, indent=2)
        return default_categories
    except:
        return default_categories

def load_data():
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except:
        pass
    return []

def save_data(data):
    try:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except:
        pass

def get_next_device_code():
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if data:
                    max_code = max(int(item[0]) for item in data)
                    return str(max_code + 1)
    except:
        pass
    return "1"

class QuanLyLoaiThietBi(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Quản Lý Loại Thiết Bị")
        self.geometry("500x500")
        self.resizable(False, False)
        self.master = master
        self.configure(bg="#FFFFFF")

        self.button_style = {
            "font": ("Arial", 11, "bold"),
            "relief": "flat",
            "padx": 10,
            "pady": 6,
            "width": 12,
            "fg": "white"
        }

        self.create_widgets()
        self.load_data()

    def create_widgets(self):
        # Header
        header_frame = tk.Frame(self, bg="#003087")
        header_frame.pack(fill=tk.X)
        tk.Label(header_frame, 
                text="📋 Danh Sách Loại Thiết Bị",
                font=("Arial", 14, "bold"),
                bg="#003087",
                fg="#FFFFFF").pack(pady=10)

        # Frame chứa danh sách
        list_frame = tk.Frame(self, bg="#FFFFFF")
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        self.listbox = tk.Listbox(list_frame, 
                                height=15, 
                                width=40, 
                                font=("Arial", 12),
                                selectmode=tk.SINGLE,
                                bg="#F0F8FF",
                                fg="#2c3e50")
        self.listbox.pack(fill=tk.BOTH, expand=True)

        # Frame chứa các nút
        button_frame = tk.Frame(self, bg="#FFFFFF")
        button_frame.pack(pady=10)

        buttons = [
            ("Thêm", self.them_loai, "#003087"),
            ("Sửa", self.sua_loai, "#003087"),
            ("Xóa", self.xoa_loai, "#FF4444")
        ]

        for text, command, color in buttons:
            btn = tk.Button(button_frame, text=text, command=command, bg=color, **self.button_style)
            btn.pack(side=tk.LEFT, padx=5)
            btn.bind("<Enter>", lambda e, b=btn, c=color: b.config(bg="#004D99" if c == "#003087" else "#FF6666"))
            btn.bind("<Leave>", lambda e, b=btn, c=color: b.config(bg=c))

    def load_data(self):
        try:
            self.listbox.delete(0, tk.END)
            categories = load_categories()
            if categories:
                for category in categories:
                    self.listbox.insert(tk.END, category)
            else:
                messagebox.showwarning("Cảnh báo", "Không có dữ liệu loại thiết bị")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tải dữ liệu: {str(e)}")

    def luu_tu_dong(self):
        try:
            categories = list(self.listbox.get(0, tk.END))
            with open(CATEGORY_FILE, "w", encoding="utf-8") as f:
                json.dump(categories, f, ensure_ascii=False, indent=2)
            if hasattr(self.master, 'lam_moi'):
                self.master.lam_moi()
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu dữ liệu: {str(e)}")

    def them_loai(self):
        dialog = tk.Toplevel(self)
        dialog.title("Thêm loại thiết bị")
        dialog.geometry("400x200")
        dialog.resizable(False, False)
        dialog.configure(bg="#FFFFFF")

        main_frame = tk.Frame(dialog, bg="#FFFFFF", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(main_frame, 
                text="Tên loại thiết bị:", 
                font=("Arial", 11),
                bg="#FFFFFF",
                fg="#2c3e50").pack(pady=10)
        
        entry = tk.Entry(main_frame, width=30, font=("Arial", 11))
        entry.pack(pady=10)

        button_frame = tk.Frame(main_frame, bg="#FFFFFF")
        button_frame.pack(pady=20)

        def save():
            new_category = entry.get().strip()
            if new_category:
                self.listbox.insert(tk.END, new_category)
                self.luu_tu_dong()
                dialog.destroy()
            else:
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập tên loại thiết bị")

        tk.Button(button_frame, 
                 text="Lưu",
                 command=save,
                 bg="#003087",
                 **self.button_style).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame,
                 text="Hủy",
                 command=dialog.destroy,
                 bg="#FF4444",
                 **self.button_style).pack(side=tk.LEFT, padx=5)

    def sua_loai(self):
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn loại thiết bị cần sửa")
            return

        dialog = tk.Toplevel(self)
        dialog.title("Sửa loại thiết bị")
        dialog.geometry("400x200")
        dialog.resizable(False, False)
        dialog.configure(bg="#FFFFFF")

        main_frame = tk.Frame(dialog, padx=20, pady=20, bg="#FFFFFF")
        main_frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(main_frame, text="Tên loại thiết bị:", font=("Arial", 11), bg="#FFFFFF").pack(pady=10)
        entry = tk.Entry(main_frame, width=30, font=("Arial", 11))
        entry.insert(0, self.listbox.get(selection[0]))
        entry.pack(pady=10)

        button_frame = tk.Frame(main_frame, bg="#FFFFFF")
        button_frame.pack(pady=20)

        def save():
            new_category = entry.get().strip()
            if new_category:
                self.listbox.delete(selection[0])
                self.listbox.insert(selection[0], new_category)
                self.luu_tu_dong()
                dialog.destroy()

        tk.Button(button_frame, text="Lưu",
                 command=save,
                 bg="#003087",
                 **self.button_style).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Hủy",
                 command=dialog.destroy,
                 bg="#FF4444",
                 **self.button_style).pack(side=tk.LEFT, padx=5)

    def xoa_loai(self):
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn loại thiết bị cần xóa")
            return

        if messagebox.askyesno("Xác nhận", "Bạn có chắc chắn muốn xóa loại thiết bị này?"):
            self.listbox.delete(selection[0])
            self.luu_tu_dong()

class NhapThongTinVatDung(tk.Toplevel):
    def __init__(self, master, on_submit, default_data=None):
        super().__init__(master)
        self.title("Nhập Thông Tin Vật Dụng")
        self.geometry("600x350")
        self.resizable(False, False)
        self.configure(bg="#FFFFFF")
        self.on_submit = on_submit
        self.image_path = tk.StringVar()

        categories = load_categories()
        if not categories:
            categories = ["Thiết bị mặc định"]
            with open(CATEGORY_FILE, "w", encoding="utf-8") as f:
                json.dump(categories, f, ensure_ascii=False, indent=2)

        self.default_data = default_data or ["", "", "Có sẵn", "", "1", categories[0]]
        
        self.main_frame = tk.Frame(self, padx=10, pady=10, bg="#FFFFFF")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        self.create_widgets()

    def create_widgets(self):
        header_frame = tk.Frame(self.main_frame, bg="#003087")
        header_frame.pack(fill=tk.X)
        tk.Label(header_frame, 
                text="📋 Nhập Thông Tin Vật Dụng",
                font=("Arial", 14, "bold"),
                bg="#003087",
                fg="#FFFFFF").pack(pady=10)

        input_frame = tk.Frame(self.main_frame, bg="#FFFFFF")
        input_frame.pack(fill=tk.X, pady=10)

        labels = ["Mã:", "Tên vật dụng:", "Trạng thái:", "Người mượn:", "Số lượng:", "Loại thiết bị:"]
        self.entries = []

        for i, label in enumerate(labels):
            label_widget = tk.Label(input_frame, text=label, font=("Arial", 11), bg="#FFFFFF", fg="#2c3e50")
            label_widget.grid(row=i, column=0, padx=(0, 10), pady=5, sticky="e")
            
            if i == 4:
                entry = tk.Entry(input_frame, width=35, validate="key", font=("Arial", 11))
                entry['validatecommand'] = (entry.register(self.validate_number), '%P')
                entry.insert(0, self.default_data[i])
            elif i == 5:
                categories = load_categories()
                entry = ttk.Combobox(input_frame, width=32, values=categories, state="readonly", font=("Arial", 11))
                if self.default_data[i]:
                    entry.set(self.default_data[i])
                else:
                    entry.set(categories[0] if categories else "")
            else:
                entry = tk.Entry(input_frame, width=35, font=("Arial", 11))
                entry.insert(0, self.default_data[i])
            
            entry.grid(row=i, column=1, pady=5, sticky="w")
            self.entries.append(entry)

        button_frame = tk.Frame(self.main_frame, bg="#FFFFFF")
        button_frame.pack(pady=10)

        save_btn = tk.Button(
            button_frame,
            text="Lưu",
            command=self.submit,
            bg="#003087",
            fg="white",
            font=("Arial", 11, "bold"),
            relief="flat",
            padx=10,
            pady=6
        )
        save_btn.pack(side=tk.LEFT, padx=5)
        save_btn.bind("<Enter>", lambda e: save_btn.config(bg="#004D99"))
        save_btn.bind("<Leave>", lambda e: save_btn.config(bg="#003087"))

    def validate_number(self, value):
        if value == "":
            return True
        try:
            int(value)
            return True
        except ValueError:
            return False

    def submit(self):
        values = [entry.get().strip() for entry in self.entries]
        if not values[0] or not values[1]:
            messagebox.showwarning("Thiếu thông tin", "Mã và tên vật dụng là bắt buộc.")
            return
        if not values[4] or int(values[4]) <= 0:
            messagebox.showwarning("Lỗi", "Số lượng phải lớn hơn 0.")
            return
        values.append(self.image_path.get())
        self.on_submit(values)
        self.destroy()

class DanhMucVatDung(tk.Toplevel):
    def __init__(self, master=None, on_change_callback=None):
        super().__init__(master)
        self.title("Danh Mục Vật Dụng")
        self.geometry("800x500")
        self.resizable(False, False)
        self.on_change_callback = on_change_callback
        self.configure(bg="#FFFFFF")

        self.button_style = {
            "font": ("Arial", 11, "bold"),
            "relief": "flat",
            "padx": 10,
            "pady": 6,
            "fg": "white"
        }
        
        self.create_widgets()
        self.lam_moi()

    def create_widgets(self):
        # Header
        header_frame = tk.Frame(self, bg="#003087")
        header_frame.pack(fill=tk.X)
        tk.Label(header_frame, text="📋 Danh Mục Vật Dụng", 
                font=("Arial", 14, "bold"),
                bg="#003087",
                fg="#FFFFFF").pack(pady=10)

        # Filter frame
        filter_frame = tk.Frame(self, bg="#FFFFFF")
        filter_frame.pack(fill=tk.X, pady=5)

        tk.Label(filter_frame, text="🔍 Tìm kiếm:", font=("Arial", 12), bg="#FFFFFF", fg="#2c3e50").pack(side="left", padx=5)
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(filter_frame, textvariable=self.search_var, width=30, font=("Arial", 12))
        search_entry.pack(side="left", padx=5)
        search_entry.bind("<KeyRelease>", lambda e: self.lam_moi(only_filter=True))

        # Frame cho bảng
        table_frame = tk.Frame(self, bg="#FFFFFF", bd=2, relief="groove")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        columns = ("Mã", "Loại thiết bị", "Tên", "Số lượng")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=10)
        self.tree.configure(selectmode="extended")
        
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        for col in columns:
            self.tree.heading(col, text=col)
        self.tree.column("Mã", width=60, anchor="center")
        self.tree.column("Loại thiết bị", width=120, anchor="center")
        self.tree.column("Tên", width=250, anchor="w")
        self.tree.column("Số lượng", width=60, anchor="center")

        # Frame cho các nút
        buttons_frame = tk.Frame(self, bg="#FFFFFF")
        buttons_frame.pack(pady=10)

        buttons = [
            ("Thêm", self.them_vat_dung, "#003087"),
            ("Sửa", self.sua_vat_dung, "#003087"),
            ("Xóa", self.xoa_vat_dung, "#003087"),
            ("Lưu", self.luu_data, "#003087"),
            ("Làm mới", self.lam_moi, "#FFC107"),
            ("Xuất Excel", self.xuat_excel, "#FFC107"),
            ("Nhập Excel", self.nhap_excel, "#FFC107"),
            ("Loại thiết bị", self.quan_ly_loai, "#FFC107")
        ]

        for text, command, color in buttons:
            btn = tk.Button(buttons_frame, text=text, command=command, bg=color, **self.button_style)
            btn.pack(side=tk.LEFT, padx=5)
            btn.bind("<Enter>", lambda e, b=btn, c=color: b.config(bg="#004D99" if c == "#003087" else "#FF8F00"))
            btn.bind("<Leave>", lambda e, b=btn, c=color: b.config(bg=c))

    def quan_ly_loai(self):
        QuanLyLoaiThietBi(self)

    def lam_moi(self, only_filter=False):
        self.tree.delete(*self.tree.get_children())
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = []

        keyword = self.search_var.get().lower()

        for item in data:
            if keyword and (keyword not in str(item[0]).lower() and keyword not in str(item[1]).lower()):
                continue
            self.tree.insert("", "end", values=(item[0], item[5] if len(item) > 5 else "", item[1], item[4]))

    def luu_data(self):
        data = []
        for item in self.tree.get_children():
            values = self.tree.item(item)["values"]
            data.append([values[0], values[2], "Có sẵn", "", values[3], values[1], ""])
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    def xuat_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Danh Muc Vat Dung"
        ws.append(["Mã", "Loại Thiết Bị", "Tên Vật Dụng", "Số Lượng", "Ảnh"])

        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = []

        for item in data:
            image_info = item[6] if len(item) > 6 else ""
            loai_thiet_bi = item[5] if len(item) > 5 else ""
            ws.append([item[0], loai_thiet_bi, item[1], item[4], image_info])

        try:
            wb.save("danh_sach_vat_dung.xlsx")
            messagebox.showinfo("Thành công", "Đã xuất file danh_sach_vat_dung.xlsx")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu file: {e}")

    def nhap_excel(self):
        file_path = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        try:
            wb = load_workbook(file_path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            for row in rows:
                if row and len(row) >= 3:
                    self.tree.insert("", "end", values=(
                        row[0],
                        row[1] if len(row) > 1 else load_categories()[0] if load_categories() else "",
                        row[2],
                        row[3] if len(row) > 3 else 1
                    ))
                    tao_ma_qr(row[0], row[2])

            self.luu_data()
            if self.on_change_callback:
                self.on_change_callback()
            messagebox.showinfo("Thành công", f"Đã nhập dữ liệu từ {os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file: {e}")

    def them_vat_dung(self):
        NhapThongTinVatDung(self, self.insert_item)

    def sua_vat_dung(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn vật dụng cần sửa")
            return

        current_values = list(self.tree.item(selection[0])["values"])
        
        edit_window = tk.Toplevel(self)
        edit_window.title("Sửa thông tin vật dụng")
        edit_window.geometry("450x300")
        edit_window.resizable(False, False)
        edit_window.configure(bg="#FFFFFF")

        header_frame = tk.Frame(edit_window, bg="#003087")
        header_frame.pack(fill=tk.X)
        tk.Label(header_frame, 
                text="📝 Sửa Thông Tin Vật Dụng",
                font=("Arial", 14, "bold"),
                bg="#003087",
                fg="#FFFFFF").pack(pady=10)

        main_frame = tk.Frame(edit_window, padx=20, pady=20, bg="#FFFFFF")
        main_frame.pack(fill=tk.BOTH, expand=True)

        input_frame = tk.Frame(main_frame, bg="#FFFFFF")
        input_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Label(input_frame, text="Tên vật dụng:", font=("Arial", 11, "bold"), bg="#FFFFFF").pack(anchor='w', pady=(0,5))
        name_entry = tk.Entry(input_frame, width=40, font=("Arial", 11))
        name_entry.insert(0, current_values[2])
        name_entry.pack(fill=tk.X, pady=(0,15))

        tk.Label(input_frame, text="Số lượng:", font=("Arial", 11, "bold"), bg="#FFFFFF").pack(anchor='w', pady=(0,5))
        quantity_entry = tk.Entry(input_frame, width=40, font=("Arial", 11))
        quantity_entry.insert(0, current_values[3])
        quantity_entry.pack(fill=tk.X, pady=(0,15))

        def validate_and_save():
            new_name = name_entry.get().strip()
            new_quantity = quantity_entry.get().strip()

            if not new_name:
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập tên vật dụng")
                return

            try:
                new_quantity = int(new_quantity)
                if new_quantity < 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("Cảnh báo", "Số lượng phải là số nguyên không âm")
                return

            current_values[2] = new_name
            current_values[3] = str(new_quantity)
            self.tree.item(selection[0], values=current_values)
            self.luu_data()
            messagebox.showinfo("Thành công", "Đã cập nhật thông tin vật dụng")
            edit_window.destroy()

        button_frame = tk.Frame(main_frame, bg="#FFFFFF")
        button_frame.pack(pady=20)

        save_btn = tk.Button(
            button_frame,
            text="Lưu thay đổi",
            command=validate_and_save,
            bg="#003087",
            fg="white",
            font=("Arial", 11, "bold"),
            relief="flat",
            padx=10,
            pady=6
        )
        save_btn.pack(side=tk.LEFT, padx=5)
        save_btn.bind("<Enter>", lambda e: save_btn.config(bg="#004D99"))
        save_btn.bind("<Leave>", lambda e: save_btn.config(bg="#003087"))

        cancel_btn = tk.Button(
            button_frame,
            text="Hủy",
            command=edit_window.destroy,
            bg="#FF4444",
            fg="white",
            font=("Arial", 11, "bold"),
            relief="flat",
            padx=10,
            pady=6
        )
        cancel_btn.pack(side=tk.LEFT, padx=5)
        cancel_btn.bind("<Enter>", lambda e: cancel_btn.config(bg="#FF6666"))
        cancel_btn.bind("<Leave>", lambda e: cancel_btn.config(bg="#FF4444"))

    def xoa_vat_dung(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn ít nhất một vật dụng để xóa.")
            return
        
        if messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa {len(selected)} vật dụng đã chọn?"):
            for item in selected:
                values = self.tree.item(item)["values"]
                ma = values[0]
                qr_path = os.path.join(QR_CODE_DIR, f"{ma}.png")
                if os.path.exists(qr_path):
                    os.remove(qr_path)
                self.tree.delete(item)
            self.luu_data()
            if self.on_change_callback:
                self.on_change_callback()

    def insert_item(self, values):
        self.tree.insert("", "end", values=(values[0], values[5] if len(values) > 5 else "", values[1], values[4]))
        tao_ma_qr(values[0], values[1])
        self.luu_data()
        if self.on_change_callback:
            self.on_change_callback()

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    app = DanhMucVatDung(root)
    root.mainloop()